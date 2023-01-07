from openpyxl import load_workbook


workbook = load_workbook(filename="C:/Users/SergeyML/Desktop/Excell/TestCoder.xlsx")

# create sheet() helps with sheet creation
workbook.create_sheet("TestSheet2")
print(workbook.sheetnames)

#  Also add a new sheet at a perticular postion
workbook.create_sheet("TestSheet3", 0)

# To access the sheet use the sheetname
test_sheet_3 = workbook["TestSheet3"]


# To copy the sheet use copy_worksheet function
# workbook.copy_worksheet(test_sheet_3)
# print(workbook.sheetnames)




# To remove the sheet pass the sheet as an argument
# workbook.remove(test_sheet_3)
# print(workbook.sheetnames)



# Save workbook
workbook.save(filename="C:/Users/SergeyML/Desktop/Excell/TestCoder.xlsx")


