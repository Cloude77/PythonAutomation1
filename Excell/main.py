import openpyxl

from openpyxl import Workbook, load_workbook
wb = load_workbook('C:/Users/SergeyML/Desktop/Excell/test.xlsx')

ws = wb.active
print(ws)

wb = openpyxl.load_workbook('C:/Users/SergeyML/Desktop/Excell/test.xlsx')
wb_sheet = wb['Code test']
wb_sheet.title = 'Code test'
wb_sheet.sheet_properties.tabColor = 'FF0000'
wb.save('C:/Users/SergeyML/Desktop/Excell/test.xlsx')