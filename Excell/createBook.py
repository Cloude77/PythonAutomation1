from openpyxl import load_workbook

# load the test workbook
workbook = load_workbook(filename="C:/Users/SergeyML/Desktop/Excell/TestCoder.xlsx")

# Note the str() function is used to convert the return type to string
print("Sheetnames are " + str(workbook.sheetnames))
sheet = workbook.active

# Access columns A and B - gives cell tuple as output
print(sheet["A:B"])

# Access row 1
print(sheet[1])

# Access sheet element value by id
print("Value of cell A1: " + sheet["A1"].value) # Value of cell A1: Names