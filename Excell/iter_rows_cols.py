from openpyxl import load_workbook

workbook = load_workbook(filename="C:/Users/SergeyML/Desktop/Excell/TestCoder.xlsx")
sheet = workbook.active

# Using iterator to get rows values

for value in sheet.iter_rows(
    min_row=1, max_row=2, min_col=1, max_col=3, values_only=True
):
    print("Rows Values:" + str(value))

for value in sheet.iter_cols(
    min_row=1, max_row=3, min_col=1, max_col=3, values_only=True
):
    print("Column Values:" + str(value))
    
print(end='==================================================== \n')

# looping for row
# looping through row and its values

for rows in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3):
    for row in rows:
        print("Row is: " + str(row) + " and value is: " + str(row.value))